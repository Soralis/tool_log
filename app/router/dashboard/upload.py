from fastapi import APIRouter, Request, File, UploadFile, Form
from fastapi.responses import HTMLResponse

from sqlmodel import Session, select
from sqlalchemy.dialects.postgresql import insert
import io
import openpyxl
import pandas as pd
from datetime import datetime as dt

from app.models import (OrderCompletion, OrderCompletionCreate, User, Workpiece, 
                        ToolConsumption, ToolConsumptionCreate, Tool, Machine, Manufacturer, 
                        ToolType, ToolOrder, ToolOrderCreate, OrderDelivery, OrderDeliveryCreate)
from app.database_config import engine
from app.templates.jinja_functions import templates

router = APIRouter(
    prefix="/upload",
    tags=["upload"]
)


async def get_excel_sheets(file: UploadFile):
    """Get list of sheets in Excel file"""
    if not file.filename.lower().endswith('.xlsx'):
        return None
    
    content = await file.read()
    await file.seek(0)  # Reset file pointer for future reads
    xlsx = io.BytesIO(content)
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    sheets = wb.sheetnames
    return sheets

async def preview_excel_sheet(file: UploadFile, sheet_name: str = None):
    """Preview first 10 rows of specified sheet without header detection"""
    if not file.filename.lower().endswith('.xlsx'):
        return None
    
    content = await file.read()
    await file.seek(0)  # Reset file pointer for future reads
    xlsx = io.BytesIO(content)
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[sheet_name]

    df = pd.DataFrame(ws.values)
    preview_rows = min(10, len(df))  # Show up to 10 rows for better header selection

    return {
        'rows': df.head(preview_rows).values.tolist(),
        'total_rows': len(df)
    }

async def read_excel(file: UploadFile, sheet_names: str = None, header_row: int = None):
    """Read specified sheets from Excel file and return combined DataFrame"""
    if not file.filename.lower().endswith('.xlsx'):
        return None
    
    content = await file.read()
    await file.seek(0)  # Reset file pointer for future reads
    xlsx = io.BytesIO(content)
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    
    # Convert comma-separated string to list, or use all sheets
    sheet_list = sheet_names.split(',') if sheet_names else wb.sheetnames
    print(f"Processing sheets: {sheet_list}")
    
    dfs = []
    for sheet_name in sheet_list:
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        df = pd.DataFrame(ws.values)
        
        if header_row is not None and header_row < len(df):
            # Use the specified row as header
            df.columns = df.iloc[header_row]
            df = df[header_row + 1:]
            print(f"Using row {header_row} as header")
            print(f"Columns: {df.columns.tolist()}")
        else:
            print("Warning: Invalid header row specified")
            continue

        df.dropna(axis=1, thresh=3, inplace=True)
        dfs.append(df)

    final_df = pd.concat(dfs) if dfs else None
    if final_df is not None:
        print(f"Successfully combined {len(dfs)} sheets, total rows: {len(final_df)}")
    else:
        print("Warning: No data frames to combine")
    return final_df

async def write_to_db(records: list, model, create_model, session: Session, result: dict):
    valid_records = []
    for record in records:
        try:
            # Validate record by attempting to create model instance
            validated_data = create_model(**record).dict(exclude_unset=True)
            valid_records.append(validated_data)
        except Exception as e:
            print(f"Data validation error: {e}")
            result['bad_data'] += 1
            result['total_records'] += 1
            continue

    if valid_records:
        stmt = insert(model).values(valid_records)
        stmt = stmt.on_conflict_do_nothing()
        stmt = stmt.returning(model.id)

        inserted_rows = session.exec(stmt).fetchall()
        num_skipped = len(valid_records) - len(inserted_rows)
        session.commit()    

        result['total_records'] += len(valid_records)
        result['inserted'] += len(inserted_rows)
        result['skipped'] += num_skipped

    return result

@router.get("/", response_class=HTMLResponse)
async def upload_page(request: Request):
    """Render the file upload page"""
    # Also return the last time data was updated
    update_dates = {}
    with Session(engine) as session:
        # Most recent tool consumption upload
        last_tool_consumption = session.exec(
            select(ToolConsumption).order_by(ToolConsumption.datetime.desc())
        ).first()
        update_dates['tool_consumption'] = last_tool_consumption.datetime if last_tool_consumption else None

        # Most recent parts production upload
        last_parts_production = session.exec(
            select(OrderCompletion).order_by(OrderCompletion.date.desc(), OrderCompletion.time.desc())
        ).first()
        if last_parts_production:
            try:
                update_dates['parts_produced'] = dt.combine(
                    last_parts_production.date, last_parts_production.time
                )
            except (AttributeError, TypeError):
                update_dates['parts_produced'] = last_parts_production.date
        else:
            update_dates['parts_produced'] = None

        # Most recent tool orders upload
        last_tool_order = session.exec(
            select(ToolOrder).order_by(ToolOrder.order_date.desc())
        ).first()
        update_dates['tool_orders'] = last_tool_order.order_date if last_tool_order else None

        # Most recent order deliveries upload
        last_order_delivery = session.exec(
            select(OrderDelivery).order_by(OrderDelivery.delivery_date.desc())
        ).first()
        update_dates['order_deliveries'] = last_order_delivery.delivery_date if last_order_delivery else None

        # Placeholder for uploads without timestamps
        update_dates['tool_inventory'] = None
        update_dates['hourly_production'] = None

    return templates.TemplateResponse(
        "dashboard/upload.html.j2",
        {
            "request": request,
            'update_dates': update_dates
        }
    )

@router.post("/preview-sheets")
async def preview_sheets(file: UploadFile = File(...)):
    """Get list of sheets in Excel file"""
    sheets = await get_excel_sheets(file)
    if sheets is None:
        return {"error": "Unsupported file type"}
    return sheets


@router.post("/preview-sheet")
async def preview_sheet(
    file: UploadFile = File(...),
    sheet_name: str = Form(None),
):
    """Preview first 10 rows of specified sheet"""
    if not sheet_name or sheet_name == "undefined":
        return {"error": "Sheet name is required"}
    
    try:
        preview = await preview_excel_sheet(file, sheet_name)
        if preview is None:
            return {"error": "Unsupported file type or invalid sheet name"}
        return preview
    except Exception as e:
        return {"error": str(e)}

@router.post("/tool-consumption")
async def upload_tool_consumption(
    file: UploadFile = File(...),
    sheet_names: str = Form(None),
    header_row: int = Form(None)
):
    """Handle tool consumption file upload

    Returns:
        dict: A dictionary containing the following keys:
            - filename (str): The name of the uploaded file.
            - type (str): The type of the upload, e.g., "tool_consumption".
            - result (dict): A dictionary with the following keys:
                - total_records (int): Total number of records processed.
                - inserted (int): Number of records successfully inserted.
                - bad_data (int): Number of records with validation errors.
                - skipped (int): Number of records skipped due to conflicts or other issues.
    """
    print(f'Processing tool consumption file: {file.filename}')

    # Validate sheet names
    valid_sheets = await get_excel_sheets(file)
    if sheet_names:
        invalid_sheets = [sheet for sheet in sheet_names.split(',') if sheet not in valid_sheets]
        if invalid_sheets:
            return {"error": f"Invalid sheet names: {', '.join(invalid_sheets)}"}
    
    df = await read_excel(file, sheet_names, header_row)

    if df is None:
        return {"filename": file.filename, "type": "tool_consumption", "error": "Failed to read file"}

    with Session(engine) as session:
        users = session.exec(select(User)).all()
        users = {user.number: user.id for user in users}

        machines = session.exec(select(Machine)).all()
        machines = {str(machine.cost_center): machine for machine in machines}

        tools = session.exec(select(Tool)).all()
        tools = {tool.number.upper(): tool for tool in tools}

        workpieces = session.exec(select(Workpiece)).all()
        workpieces = {workpiece.description: workpiece.id for workpiece in workpieces}
        
        manufacturer = session.exec(select(Manufacturer).where(Manufacturer.name=='Undefined')).first()
        tool_type = session.exec(select(ToolType).where(ToolType.name=='Undefined')).first()

        # Prepare all valid records
        records = []
        result = {'total_records': 0, 'inserted': 0, 'bad_data': 0, 'skipped': 0}
        for _, row in df.iterrows():
            try:
                user_id = None
                if row.get('Employee'):
                    user_number = row['Employee'].split('-')[1]
                    user_id = users.get(user_number, None)
                
                machine, machine_id = None, None
                if row.get('Cost Center'):
                    cost_center = str(row['Cost Center'])
                    machine = machines[cost_center]
                    machine_id = machine.id

                tool = tools.get(str(row['CPN']).strip("\n").upper())
                if tool is None:
                    new_tool = Tool(
                        number=row['CPN'].upper(),
                        # manufacturer_name=row['Description'], # not in the data yet
                        name=row['Desc1'],
                        manufacturer_id=manufacturer.id,
                        tool_type_id=tool_type.id
                    )
                    try:
                        session.add(new_tool)
                        session.commit()
                        session.refresh(new_tool)
                        tools[new_tool.number] = new_tool
                    except Exception as e:
                        print(e)
                        tool = None
                        session.rollback()
                                        
                workpiece_id = None
                if row.get('Product'):
                    workpiece_id = workpieces.get(row['Product'], None)

                recipe_id, tool_position_id = None, None

                if machine and tool and workpiece_id:
                    for recipe in machine.recipes:
                        if recipe.workpiece_id == workpiece_id:
                            recipe_id = recipe.id
                            for tool_position in recipe.tool_positions:
                                if tool_position.tool_id == tool.id:
                                    tool_position_id = tool_position.id
                                    break
                            if recipe_id:
                                break

                records.append({
                    'datetime': row['TransDate'],
                    'number': row['TransactionId'] if row.get('TransactionId') else None,
                    'consumption_type': "ISSUE", # row['Transaction Type'], # hardcoded for now untill data contains transaction type
                    'quantity': row['Qty'],
                    'value': float(row['Ext Value']) if row.get('Ext Value') and row['Ext Value'] > 0 else tool.price * row['Qty'],
                    'price': float(row['Ext Value']) / row["Qty"] if row.get('Ext Value') and row['Ext Value'] > 0 else tool.price,
                    'user_id': user_id,
                    'machine_id': machine_id,
                    'tool_id': tool.id,
                    'recipe_id': recipe_id,
                    'tool_position_id': tool_position_id,
                    'workpiece_id': workpiece_id,
                })
                tool.price = float(row['Ext Value']) / row["Qty"] if row.get('Ext Value') and row['Ext Value'] > 0 else tool.price
                if not tool.inventory:
                    tool.inventory = 0
                tool.inventory -= row['Qty']
            except Exception as e:
                print(e)

            if len(records) > 100:
                result = await write_to_db(records, ToolConsumption, ToolConsumptionCreate, session, result)
                records = []
        
        # Insert all records in a single operation, ignoring duplicates
        if records:
            result = await write_to_db(records, ToolConsumption, ToolConsumptionCreate, session, result)

        session.commit()
        

    return {"filename": file.filename, "type": "tool_consumption", 'result': result}


@router.post("/parts-produced")
async def upload_parts_produced(
    file: UploadFile = File(...),
    sheet_names: str = Form(None),
    header_row: int = Form(None)
):
    """Handle parts production file upload"""
    print(f'Processing parts production file: {file.filename}')

    df = await read_excel(file, sheet_names, header_row)

    if df is None:
        return {"filename": file.filename, "type": "parts_produced", "error": "Unsupported file type"}
        
    with Session(engine) as session:
        # users = session.exec(select(User)).all()
        workpieces = session.exec(select(Workpiece)).all()
        workpieces = {workpiece.material: workpiece.id for workpiece in workpieces}
        
        # Prepare all valid records
        records = []
        result = {'total_records': 0, 'inserted': 0, 'bad_data': 0, 'skipped': 0}
        for _, row in df.iterrows():
            try:
                records.append({
                    'quantity': row['Qty in unit of entry'],
                    'vendor': row['Vendor'],
                    'batch': row['Batch'],
                    'customer': row['Customer'],
                    'order': row['Order'],
                    'document_number': row['Material Document'],
                    'date': row['Document Date'],
                    'time': row['Time of Entry'],
                    'value': row['Amt.in loc.cur.'],
                    'workpiece_id': workpieces[row['Material']],
                })
            except Exception as e:
                print(e)

            if len(records) > 100:
                result = await write_to_db(records, OrderCompletion, OrderCompletionCreate, session, result)
                records = []
        
        # Insert all records in a single operation, ignoring duplicates
        if records:
            result = await write_to_db(records, OrderCompletion, OrderCompletionCreate, session, result)

    return {"filename": file.filename, "type": "parts_produced", 'result': result}


@router.post("/tool-orders")
async def upload_tool_orders(
    file: UploadFile = File(...),
    sheet_names: str = Form(None),
    header_row: int = Form(None)
):
    """Handle tool orders file upload"""
    print(f'Processing tool orders file: {file.filename}')

    df = await read_excel(file, sheet_names, header_row)

    if df is None:
        return {"filename": file.filename, "type": "tool_order", "error": "Failed to read file"}

    with Session(engine) as session:
        tools = session.exec(select(Tool)).all()
        tools = {tool.number.upper(): tool.id for tool in tools}
        
        manufacturers = session.exec(select(Manufacturer)).all()
        manufacturers = {int(manufacturer.number): manufacturer.id for manufacturer in manufacturers if manufacturer.number}

        unknown_tool_type = session.exec(select(ToolType).where(ToolType.name=='Undefined')).first()
        
        # Prepare all valid records
        records = []
        result = {'total_records': 0, 'inserted': 0, 'bad_data': 0, 'skipped': 0}

        for _, row in df.iterrows():
            if not row.get('OrderQty', False):
                continue
            try:
                tool_id = tools.get(str(row.get('custpart', '')).upper())
                manufacturer_id = manufacturers.get(row.get('VendorNumber'))

                if not manufacturer_id:
                    new_manufacturer = Manufacturer(
                        name=row['VendorName'],
                        number=row['VendorNumber'],
                    )
                    try:
                        session.add(new_manufacturer)
                        session.commit()
                        session.refresh(new_manufacturer)
                        manufacturer_id = new_manufacturer.id
                        manufacturers[int(new_manufacturer.number)] = manufacturer_id
                    except Exception as e:
                        print(e)
                        manufacturer_id = manufacturers['000000'] # default value if manufacturer is not found
                        result['bad_data'] += 1
                        session.rollback()

                if not tool_id:
                    new_tool = Tool(
                        number=str(row['custpart']).upper(),
                        manufacturer_name=f"{row['c_description']} - {row['Textbox18']}",
                        name=f"{row['c_description']} - {row['Textbox18']}",
                        manufacturer_id=manufacturer_id,
                        tool_type_id=unknown_tool_type.id
                    )
                    try:
                        session.add(new_tool)
                        session.commit()
                        session.refresh(new_tool)
                        tool_id = new_tool.id
                        tools[new_tool.number] = tool_id
                    except Exception as e:
                        print(e)
                        tool_id = None
                        result['bad_data'] += 1
                        session.rollback()
                        continue

                records.append({
                    'tool_id': tool_id,
                    'quantity': row['OrderQty'],
                    'number': str(row['PONumber']),
                    'suffix': str(row['POSuffix']),
                    'line': str(row['POLine']),
                    'order_date': row['OrderDate'],
                    'estimated_delivery_date': row['DueDate'],
                    'tool_price': float(row.get('OpenCost', 0)),
                    'gross_price': float(row.get('TotalCost', 0)),
                })
            except Exception as e:
                print(e)

            if len(records) > 100:
                result = await write_to_db(records, ToolOrder, ToolOrderCreate, session, result)
                records = []
        
        # Insert all records in a single operation, ignoring duplicates
        if records:
            result = await write_to_db(records, ToolOrder, ToolOrderCreate, session, result)

    return {"filename": file.filename, "type": "tool_consumption", 'result': result}


@router.post("/tool-delivery")
async def upload_tool_deliveries(
    file: UploadFile = File(...),
    sheet_names: str = Form(None),
    header_row: int = Form(None)
):
    """Handle tool deliveries file upload"""
    print(f'Processing tool deliveries file: {file.filename}')

    df = await read_excel(file, sheet_names, header_row)

    if df is None:
        return {"filename": file.filename, "type": "tool_delivery", "error": "Failed to read file"}

    with Session(engine) as session:
        orders = session.exec(select(ToolOrder)).all()
        orders = {f'{order.number}-{order.suffix}-{order.line}': order.id for order in orders}
        
        tools = session.exec(select(Tool)).all()
        tools = {tool.number: tool.id for tool in tools}

        unknown_manufacturer = session.exec(select(Manufacturer).where(Manufacturer.name == 'Undefined')).first()
        unknown_tool_type = session.exec(select(ToolType).where(ToolType.name == 'Undefined')).first()

        # Prepare all valid records
        records = []
        result = {'total_records': 0, 'inserted': 0, 'bad_data': 0, 'skipped': 0}

        for _, row in df.iterrows():
            try:
                order_id = orders.get(f"{row.get('PONumber', '')}-{row.get('POLine', '')}")

                if not order_id:
                    try:
                        tool_id=tools.get(str(row['custpart']).upper())
                        if not tool_id:
                            new_tool = Tool(
                                name=f"{row['PartDescription2']} - {row['PartDescription1']}",
                                number=str(row['custpart']).upper(),
                                tool_type_id=unknown_tool_type.id,
                                manufacturer_id=unknown_manufacturer.id
                            )
                            try:
                                session.add(new_tool)
                                session.commit()
                                session.refresh(new_tool)
                                tool_id = new_tool.id
                                tools[new_tool.number] = tool_id
                            except Exception as e:
                                result['bad_data'] += 1
                                session.rollback()
                                continue
                    
                        new_order = ToolOrder(
                            tool_id=tool_id,
                            number=row['PONumber'].split('-')[0],
                            suffix=row['PONumber'].split('-')[1],
                            line=str(row['POLine']),
                            quantity=row['OrderQty'],
                            order_date=row['OrderDate'],
                            estimated_delivery_date=row['DueDate'],
                        )
                        session.add(new_order)
                        session.commit()
                        session.refresh(new_order)
                        order_id = new_order.id
                        orders[f"{new_order.number}-{new_order.suffix}-{new_order.line}"] = order_id
                    except Exception as e:
                        print(e)
                        result['bad_data'] += 1
                        session.rollback()
                        continue

                records.append({
                    'order_id': order_id,
                    'quantity': row['ReceivedQty'],
                    'delivery_date': row['ReceiptDate'],
                })
            except Exception as e:
                print(e)

            if len(records) > 100:
                result = await write_to_db(records, OrderDelivery, OrderDeliveryCreate, session, result)
                records = []

        
        # Insert all records in a single operation, ignoring duplicates
        if records:
            result = await write_to_db(records, OrderDelivery, OrderDeliveryCreate, session, result)

    return {"filename": file.filename, "type": "tool_consumption", 'result': result}


@router.post("/tool-inventory")
async def upload_tool_inventory(
    file: UploadFile = File(...),
    sheet_names: str = Form(None),
    header_row: int = Form(None)
):
    """Handle tool inventory file upload"""
    print(f'Processing tool inventory file: {file.filename}')

    df = await read_excel(file, sheet_names, header_row)

    if df is None:
        return {"filename": file.filename, "type": "tool_delivery", "error": "Failed to read file"}

    with Session(engine) as session:      
        tools = session.exec(select(Tool)).all()
        tools = {tool.number: tool for tool in tools}

        # Prepare all valid records
        result = {'total_records': 0, 'inserted': 0, 'bad_data': 0, 'skipped': 0}

        for _, row in df.iterrows():
            tool = tools.get(row.get('Part Number', ''))

            if not tool:
                result['bad_data'] += 1
                continue

            tool.inventory = row.get('Total', 0)
            result['inserted'] += 1
            result['total_records'] += 1
            
        session.commit()

    return {"filename": file.filename, "type": "tool_consumption", 'result': result}


@router.post("/hourlyProduction")
async def upload_production(
    file: UploadFile = File(...),
    sheet_names: str = Form(None),
):
    try:
        content = await file.read()
        await file.seek(0)  # Reset file pointer for future reads
        xlsx = io.BytesIO(content)
        wb = openpyxl.load_workbook(xlsx, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return {"filename": file.filename, "type": "production", "error": "Failed to read file"}
    
    # Read the selected sheet and find inners and outers
    ws = wb[sheet_names]
    locations = {}
    workpiece_type = None
    start, end = None, None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.lower() in ['inner', 'outer']:
                workpiece_type = cell.value.lower()
                start = (cell.column, cell.row + 1)
            if start:
                if cell.column < start[0]:
                    continue
                if isinstance(cell.value, str) and any(ordinal in cell.value.lower() for ordinal in ['1st', '2nd', '3rd']):
                    end = (cell.column, cell.row - 1)
                    break
            if start and end:
                locations[workpiece_type] = {
                    'start': start,
                    'end': end
                }
                start, end, workpiece_type = None, None, None
            
    # clean up the headers and create dataframes
    for workpiece_type, loc in locations.items():
        workpiece_values = []
        for row in ws.iter_rows(min_row=loc['start'][1], max_row=loc['end'][1], min_col=loc['start'][0], max_col=loc['end'][0]):
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            workpiece_values.append(row_values)
        line_1 = workpiece_values.pop(0)
        for i in range(len(workpiece_values[0])):
            if workpiece_values[0][i] is None and line_1[i] is not None:
                workpiece_values[0][i] = line_1[i]
        locations[workpiece_type] = pd.DataFrame(workpiece_values[1:], columns=workpiece_values[0])

    # Save data from dataframes in Database
    with Session(engine) as session:
        pass # TODO: implement saving logic

    result = {'total_records': 0, 'inserted': 0, 'bad_data': 0, 'skipped': 0}
    return {"filename": file.filename, "type": "tool_consumption", 'result': result}
